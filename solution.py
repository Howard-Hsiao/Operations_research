#!/usr/bin/env python
# coding: utf-8

# In[3]:


import pandas as pd
from gurobipy import *
import numpy as np
import openpyxl

# In[98]:

CSRs_num=40

# File name
dataFile_path = "./data/OR108-2_case02_data.xlsx"

# read demandInfo
monthDay = 31
demandInfo = pd.read_excel(dataFile_path, sheet_name='demand',
                           header = None, skiprows = 3)
demandInfo.columns = ["Time"] + [i for i in range(1, monthDay+1)]
demandInfo.set_index("Time", inplace = True)

# read shift
shiftInfo = pd.read_excel(dataFile_path, sheet_name='shifts',
                           header = None, skiprows = 2)
shiftInfo.drop([0, 1], axis = 1, inplace = True)
shiftInfo.columns = ["Shift"] + list(demandInfo.index)
shiftInfo.set_index("Shift", inplace = True)


# read shift request
shiftRequest = pd.read_excel(dataFile_path, sheet_name='shift requests')


# read leave request
leaveRequest = pd.read_excel(dataFile_path, sheet_name='leave requests')


# In[104]:


# define parameter
periodNum = 24
dateNum = len(demandInfo.columns)


# In[105]:
demandInfo = demandInfo.values.tolist()
demandInfo = np.array(demandInfo)
demandInfo = np.transpose(demandInfo)
shiftInfo=shiftInfo.values.tolist()
shiftInfo = np.array(shiftInfo)

m = Model("solution")
w = list()
x = list()
# set variables
for i in range(dateNum):
    x.append(list())
    for j in range(0,14,1):
        x[i].append(m.addVar(lb = 0, vtype = GRB.CONTINUOUS, name = "x_{}_{}".format(i, j)))
x=np.array(x)

for i in range(dateNum):
    w.append(list())
    for j in range(periodNum):
        w[i].append(m.addVar(lb = 0, vtype = GRB.CONTINUOUS, name = "w_{}_{}".format(i, j)))      
m.update()

# set objective function
objFunc = gurobipy.LinExpr()
for i in range(dateNum):
    for j in range(periodNum):
        objFunc += w[i][j]
m.setObjective(objFunc, GRB.MINIMIZE)

# ----- Add in CSRs constraint -----V
for date in  range(0,31,1):
    CSRs_limits=gurobipy.LinExpr()
    for shift in x[date]:
        CSRs_limits+=shift
    m.addConstr(lhs=CSRs_limits,sense=GRB.EQUAL,rhs=CSRs_num)
        

# ----- Add in Positive Definite Constraint ----

supply = demandInfo - np.dot(x, shiftInfo)
for date in range(dateNum):
    for period in range(periodNum):
        positive_definite = gurobipy.LinExpr()
        positive_definite=w[date][period]-supply[date][period]
        m.addConstr(lhs= positive_definite, sense = GRB.GREATER_EQUAL,rhs=0)
        

# ------ Add in Month day off Constraint -------V
day_off_month_limits=gurobipy.LinExpr()
for date in range(dateNum):
    day_off_month_limits += x[date][13]
m.addConstr(lhs=day_off_month_limits,sense=GRB.GREATER_EQUAL,rhs=8*CSRs_num)


# ------- Add in Request shift -------

request_shift=shiftRequest.values.tolist()

for request in request_shift:
    request_date = int(request[3].split("/")[1])-1
    constra=gurobipy.LinExpr()
    constra=x[request_date][request[4]-1]
    m.addConstr(lhs=constra,sense=GRB.GREATER_EQUAL,rhs=1)
    

# --------- Add in Leave request -------V
leave_request_shift=leaveRequest.values.tolist()
reserve_dayoff_data=np.zeros(31)
for request in leave_request_shift:
    date=str(request[3])
    if date.count("-")>0:
        date=date.split("-")
        start=int(date[0].split("/")[1])
        end=int(date[1].split("/")[1])
        for i in range(start-1,end):
            reserve_dayoff_data[i]+=1
    else:
        start=int(date.split("/")[1])-1
        reserve_dayoff_data[start]+=1
for i in range(0,31,1):
    if reserve_dayoff_data[i]!=0:
        constra=gurobipy.LinExpr()
        constra=x[i][13]
        m.addConstr(lhs=constra,sense=GRB.GREATER_EQUAL,rhs=reserve_dayoff_data[i])

# --------- Week on shifts-------
for start in range(0,25):
    # Day off limitsV
    dayoff_shift_limits=gurobipy.LinExpr()
    for date in range(start,start+7):
        dayoff_shift_limits += x[date][13]
    m.addConstr(lhs=dayoff_shift_limits, sense=GRB.GREATER_EQUAL , rhs=CSRs_num*1)
    
    # Night limitsV
    night_shift_limits=gurobipy.LinExpr()
    for date in range(start,start+7):
        for shift in range(10,13):
            night_shift_limits += x[date][shift]
    m.addConstr(lhs=night_shift_limits, sense=GRB.LESS_EQUAL , rhs=CSRs_num*1)
    
    # Afternoon limitsV
    Afternoon_shift_limits=gurobipy.LinExpr()
    for date in range(start,start+7):
        for shift in range(6,10):
            Afternoon_shift_limits += x[date][shift]
    m.addConstr(lhs=Afternoon_shift_limits, sense=GRB.LESS_EQUAL , rhs=CSRs_num*2)
        
m.optimize()
workbook = openpyxl.load_workbook(dataFile_path)       
try:
    worksheet = workbook.get_sheet_by_name('Output Data')
    workbook.remove_sheet(worksheet)
except:
    pass
workbook.create_sheet('Output Data')
worksheet = workbook.get_sheet_by_name('Output Data')
headline=["Date"]
for num in range(1,14,1):
    headline.append(num)
headline.append(0)
worksheet.append(headline)
now_date=1
Out_data=[]
for rows in x:
    data=[]
    for value in rows:
        data.append(value.x)
    Out_data.append(data)
    worksheet.append([now_date]+data)
    now_date+=1
    
    
try:
    worksheet = workbook.get_sheet_by_name('Omega')
    workbook.remove_sheet(worksheet)
except:
    pass
workbook.create_sheet('Omega')
worksheet = workbook.get_sheet_by_name('Omega')
headline=["Date"]
for num in range(1,25,1):
    headline.append(num)
worksheet.append(headline)
now_date=1
for rows in w:
    data=[]
    for value in rows:
        data.append(value.x)
    worksheet.append([now_date]+data)
    now_date+=1

Out_data=np.array(Out_data)
Supply=demandInfo-np.dot(Out_data,shiftInfo)

try:
    worksheet = workbook.get_sheet_by_name('Output Demand')
    workbook.remove_sheet(worksheet)
except:
    pass
workbook.create_sheet('Output Demand')
worksheet = workbook.get_sheet_by_name('Output Demand')
headline=["Date"]
for num in range(1,25,1):
    headline.append(num)
now_date=1
worksheet.append(headline)
for rows in Supply:
    data = []
    for num in rows:
        data.append(num)
    worksheet.append([now_date]+data)
    now_date+=1
workbook.save(dataFile_path)
        
        
# set constraints
#m.addConstr(lhs = unitMade, sense = GRB.LESS_EQUAL, 
#            rhs = agentInfo["hourMax"]/agentInfo["hourNeed"])

